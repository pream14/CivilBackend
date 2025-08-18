from flask_jwt_extended import JWTManager, create_access_token, get_jwt_identity, jwt_required
from flask import Flask, request, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from datetime import date, datetime, timedelta  # Use this for datetime functionality
from sqlalchemy import func
from decimal import Decimal
from werkzeug.security import generate_password_hash, check_password_hash  # Import from werkzeug
from flask_cors import CORS
from sqlalchemy import or_
import joblib
import numpy as np
import pandas as pd
from sqlalchemy import func, case 
import io
import os
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

app = Flask(__name__)
CORS(app)  # This will allow all domains to access your Flask app
app.config['JWT_SECRET_KEY'] = "Jackdog02#"
app.config['JWT_ACCESS_TOKEN_EXPIRES'] =timedelta(days=365)

app.config['SQLALCHEMY_DATABASE_URI'] = 'mysql+mysqlconnector://root:Jackdog02#@localhost/Civil'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = True
db = SQLAlchemy(app)
jwt = JWTManager(app)
# Load the saved model
model = joblib.load("final_cost_prediction_model.pkl")



class options(db.Model):
    __tablename__ = 'options'
    sno = db.Column(db.Integer, primary_key=True, autoincrement=True)

    labour = db.Column(db.String(100),  nullable=False)
    machinery = db.Column(db.String(255), nullable=False)
    material = db.Column(db.String(255), nullable=False)
    


    def __init__(self,labour, material, machinery):
        self.labour=labour
        self.machinery = machinery
        self.material = material

class category(db.Model):
    __tablename__ = 'category'
    id = db.Column(db.String(15), primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    type = db.Column(db.String(255), nullable=False)


    def __init__(self,id, name, type):
        self.id=id
        self.name = name
        self.type = type

class projects(db.Model):
    __tablename__ = 'projects'
    projectname = db.Column(db.String(100), unique=True, primary_key=True)
    quotedamount = db.Column(db.BigInteger, nullable=False, default=0)
    totexpense = db.Column(db.BigInteger, nullable=False)
    startdate = db.Column(db.Date, nullable=False)  # Add this line
    duration = db.Column(db.Integer, nullable=False)  # Add this line

    def __init__(self, projectname, quotedamount, totexpense, startdate, duration):
        self.projectname = projectname
        self.quotedamount = quotedamount
        self.totexpense = totexpense
        self.startdate = startdate  # Add this line
        self.duration = duration  # Add this line


class payments1(db.Model):
    __tablename__ = 'payments1'
    pid = db.Column(db.Integer, primary_key=True, autoincrement=True)
    projectname = db.Column(db.String(100),db.ForeignKey('projects.projectname'))
    type = db.Column(db.String(100),nullable=False)
    estamount = db.Column(db.Integer, nullable=False)
    expense = db.Column(db.Integer, nullable=False)



    def __init__(self,projectname, estamount,type, expense):
        self.projectname=projectname
        self.expense = expense
        self.estamount = estamount
        self.type = type

class payments2(db.Model):
    __tablename__ = 'payments2'
    uid = db.Column(db.Integer, primary_key=True, autoincrement=True)
    projectname = db.Column(db.String(100),db.ForeignKey('projects.projectname'),nullable=False)
    id = db.Column(db.String(15), db.ForeignKey('category.id'),nullable=False)
    amount = db.Column(db.Integer, nullable=False)
    date = db.Column(db.Date, nullable=False)  

    def __init__(self,projectname, amount,id, date):
        self.projectname=projectname
        self.date = date
        self.amount = amount
        self.id = id
@app.route('/get-options', methods=['GET'])
def get_options():
    try:
        Options = options.query.all()  # ✅ Use "Options"
        options_list = [
            {
                "labour": option.labour,
                "machinery": option.machinery,
                "material": option.material
            }
            for option in Options
        ]
        return jsonify(options_list)
    except Exception as e:
        import traceback
        print("Options fetch error:", str(e))
        print(traceback.format_exc())  # ✅ Prints detailed error log
        return jsonify({"error": "Failed to fetch options", "details": str(e)}), 500

@app.route('/add-category', methods=['POST'])
def add_category():
    data = request.get_json()  
    if not data or 'name' not in data or 'type' not in data or 'id' not in data:
        return jsonify({"error": "Missing 'name', 'type', or 'id' in request data"}), 400  # Include 'error' keyword in the message
    
    id = data['id']
    name = data['name']
    type_ = data['type']

    try:
        new_category = category(name=name, type=type_, id=id)
        db.session.add(new_category)
        db.session.commit()
        return jsonify({"success": True, "message": "Category added successfully!", "id": new_category.id}), 201  # Include 'success' keyword
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": "Failed to add category. " + str(e)}), 500  # Include 'error' keyword

@app.route('/get_category', methods=['GET'])
def get_category():
    categories = category.query.all()
    return jsonify([{'id': cat.id, 'name': cat.name, 'type': cat.type} for cat in categories])

# Update category (including ID change)
@app.route('/update-category', methods=['POST'])
def update_category():
    data = request.json
    Category = category.query.filter_by(id=data['old_id']).first()  

    if Category:
        # Check if ID was changed
        if data['id'] != data['old_id']:
            existing = category.query.filter_by(id=data['id']).first()
            if existing:
                return jsonify({"error": "ID already exists!"}), 400  # Prevent duplicate ID

            Category.id = data['id']  # Change ID

        Category.name = data['name']
        Category.type = data['type']
        db.session.commit()
        return jsonify({"message": "Category updated successfully!"})
    
    return jsonify({"error": "Category not found!"}), 404

# Delete category
@app.route('/delete-category/<int:category_id>', methods=['DELETE'])
def delete_category(category_id):
    Category = category.query.filter_by(id=category_id).first()

    if Category:
        db.session.delete(Category)
        db.session.commit()
        return jsonify({"message": "Category deleted successfully!"})
    
    return jsonify({"error": "Category not found!"}), 404

@app.route('/labour', methods=['GET'])
def get_labour():
    # Query the labour column
    labour_data = db.session.query(options.labour).all()
    # Convert tuples to a list of strings
    labour_list = [labour[0] for labour in labour_data]
    # Return as JSON
    print(labour_list)
    return jsonify({'labour': labour_list})

@app.route('/add-labour', methods=['POST'])
def add_labour():
    data = request.get_json()
    
    # Validate the incoming data
    if not data or 'labour' not in data:
        return jsonify({"error": "Missing 'labour' field in request data"}), 400  # Include 'error' keyword
    
    labour_type = data['labour']
    
    try:
        # Ensure labour_type is unique (example query, adjust for your DB structure)
        existing_labour = db.session.query(options).filter_by(labour=labour_type).first()
        if existing_labour:
            return jsonify({"error": "Labour type already exists."}), 400  # Prevent duplicate additions

        # Add new labour type
        new_labour = options(labour=labour_type, machinery='', material='')  # Default values for machinery and material
        db.session.add(new_labour)
        db.session.commit()
        
        return jsonify({"success": True, "message": "Labour added successfully!", "labour": labour_type}), 201  # Include 'success' keyword
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": "Failed to add labour. " + str(e)}), 500  # Include 'error' keyword

@app.route('/machinery', methods=['GET'])
def get_machinery():
    # Query the labour column
    machinery_data = db.session.query(options.machinery).all()
    # Convert tuples to a list of strings
    machinery_list = [machinery[0] for machinery in machinery_data]
    # Return as JSON
    print(machinery_list)
    return jsonify({'machinery': machinery_list})


@app.route('/add-machinery', methods=['POST'])
def add_machinery():
    data = request.get_json()
    
    # Validate the incoming data
    if not data or 'machinery' not in data:
        return jsonify({"error": "Missing 'machinery' field in request data"}), 400  # Include 'error' keyword
    
    machinery_type = data['machinery']
    
    try:
        # Ensure labour_type is unique (example query, adjust for your DB structure)
        existing_machinery = db.session.query(options).filter_by(machinery=machinery_type).first()
        if existing_machinery:
            return jsonify({"error": "Machinery type already exists."}), 400  # Prevent duplicate additions

        # Add new labour type
        new_machinery = options(labour='', machinery=machinery_type, material='')  # Default values for machinery and material
        db.session.add(new_machinery)
        db.session.commit()
        
        return jsonify({"success": True, "message": "Machinery added successfully!", "machinery": machinery_type}), 201  # Include 'success' keyword
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": "Failed to add machinery. " + str(e)}), 500  # Include 'error' keyword
    

@app.route('/material', methods=['GET'])
def get_material():
    # Query the labour column
    material_data = db.session.query(options.material).all()
    # Convert tuples to a list of strings
    material_list = [material[0] for material in material_data]
    # Return as JSON
    print(material_list)
    return jsonify({'material': material_list})


@app.route('/add-material', methods=['POST'])
def add_material():
    data = request.get_json()
    
    # Validate the incoming data
    if not data or 'material' not in data:
        return jsonify({"error": "Missing 'material' field in request data"}), 400  # Include 'error' keyword
    
    material_type = data['material']
    
    try:
        # Ensure labour_type is unique (example query, adjust for your DB structure)
        existing_material = db.session.query(options).filter_by(material=material_type).first()
        if existing_material:
            return jsonify({"error": "Material type already exists."}), 400  # Prevent duplicate additions

        # Add new labour type
        new_material = options(labour='', machinery='', material=material_type)  # Default values for machinery and material
        db.session.add(new_material)
        db.session.commit()
        
        return jsonify({"success": True, "message": "Material added successfully!", "material": material_type}), 201  # Include 'success' keyword
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": "Failed to add material. " + str(e)}), 500  # Include 'error' keyword




@app.route('/submit_project', methods=['POST'])
def submit_project():
    data = request.json
    project_name = data.get('projectname')
    estimated_amount = data.get('estimated_amount', 0)  # Fixed name
    start_date = data.get('start_date', datetime.today().strftime('%Y-%m-%d'))  # Use current date if not provided
    duration = data.get('duration', 0)  # Fixed name
    payments = data.get('rows', [])  # Fixed from 'payments' to 'rows'

    # Check if project already exists
    existing_project = projects.query.filter_by(projectname=project_name).first()

    if not existing_project:
        # Insert project into projects table
        new_project = projects(projectname=project_name, quotedamount=estimated_amount, totexpense=0,startdate=start_date,duration=duration)
        db.session.add(new_project)
        db.session.commit()  # Commit immediately to avoid foreign key issues

    # Insert into payments1
    for payment in payments:
        new_payment = payments1(
            projectname=project_name,
            type=payment.get('type'),
            estamount=payment.get('estamount'),
            expense=payment.get('expense', 0)
        )
        db.session.add(new_payment)

    db.session.commit()  # Commit all payments

    return jsonify({"message": "Project and payments added successfully"}), 200

@app.route('/add_expense', methods=['POST'])
def add_expense():
    data = request.get_json()
    projectname = data.get('projectname')
    category_id = data.get('category_id')  # ID from category table
    expense_amount = data.get('expense')
    selected_type = data.get('type')  # Labour, Material, Machinery
    date = data.get('date', datetime.today().strftime('%Y-%m-%d'))  # Use current date if not provided

    if not projectname or not category_id or not expense_amount or not selected_type:
        return jsonify({"error": "Missing required fields"}), 400

    try:
        # Fetch the project and update total expense
        project = projects.query.filter_by(projectname=projectname).first()
        if not project:
            return jsonify({"error": "Project not found"}), 404

        project.totexpense += int(expense_amount)

        # Update the expense in payments1
        payment_entry = payments1.query.filter_by(projectname=projectname, type=selected_type).first()
        if payment_entry:
            payment_entry.expense += int(expense_amount)
        else:
            return jsonify({"error": "Payment entry not found for given project and type"}), 404

        # Insert new row into payments2
        new_payment = payments2(
            projectname=projectname,
            id=category_id,
            amount=int(expense_amount),
            date=datetime.strptime(date, '%Y-%m-%d')
        )
        db.session.add(new_payment)

        # Commit changes to the database
        db.session.commit()

        return jsonify({"message": "Expense added successfully"}), 200

    except Exception as e:
        db.session.rollback()  # Rollback in case of error
        return jsonify({"error": str(e)}), 500


@app.route('/get_project_payment_details', methods=['GET'])
def get_project_payment_details():
    projectname = request.args.get('projectname')
    selected_type = request.args.get('type')  # Labour, Material, Machinery

    if not projectname or not selected_type:
        return jsonify({"error": "Missing projectname or type"}), 400

    # Check if the project exists in payments1 for the given type
    project_exists = (
        db.session.query(payments1)
        .filter(payments1.projectname == projectname, payments1.type == selected_type)
        .first()
    )

    if not project_exists:
        return jsonify({"error": "add the quoted amount for selected typr"+selected_type}), 404

    # Fetch all category names and IDs where type matches
    results = (
        db.session.query(category.id.label('category_id'), category.name.label('category_name'))
        .filter(category.type == selected_type)
        .all()
    )

    if results:
        categories = [{"category_id": row.category_id, "category_name": row.category_name} for row in results]
        print(categories)
        return jsonify({
            "projectname": projectname,
            "categories": categories
        })
    else:
        return jsonify({"error": "No matching categories found"}), 404


@app.route('/search', methods=['GET'])
def search_projects():
    search_text = request.args.get('project_name', '').strip()

    if not search_text:
        return jsonify({"filtered_projects": []})  # Return empty list if no input

    try:
        # Query database for matching projects
        project_list = projects.query.filter(projects.projectname.ilike(f"%{search_text}%")).all()

        # Convert results to a list of dictionaries
        filtered_projects = [{"projectname": project.projectname} for project in project_list]

        return jsonify({"filtered_projects": filtered_projects})

    except Exception as e:
        print(f"Error in /search: {str(e)}")  # Log error to console
        return jsonify({"error": "Internal server error"}), 500  # Return a proper error response


@app.route('/projects', methods=['GET'])
def get_projects():
    projects_list = projects.query.all()
    projects_data = [
        {
            "projectname": proj.projectname,
            "quotedamount": proj.quotedamount,
            "totexpense": proj.totexpense
        }
        for proj in projects_list
    ]
    return jsonify({"projects": projects_data})


@app.route('/project-details', methods=['GET'])
def get_project_details():
    projectname = request.args.get('projectname')
    if not projectname:
        return jsonify({"error": "Missing projectname"}), 400

    project = projects.query.filter_by(projectname=projectname).first()
    if not project:
        return jsonify({"error": "Project not found"}), 404

    return jsonify({
        "projectname": project.projectname,
        "quotedamount": project.quotedamount,
        "totexpense": project.totexpense
    }), 200

@app.route('/employee-details', methods=['GET'])
def get_employee_details():
    projectname = request.args.get('projectname')
    if not projectname:
        return jsonify({"error": "Missing projectname"}), 400

    employee_data = (
        db.session.query(category.name, category.type, payments2.date, payments2.amount)
        .join(payments2, category.id == payments2.id)
        .filter(payments2.projectname == projectname)
        .all()
    )

    result = [
        {
            "name": row.name,
            "type": row.type,
            "date": row.date.strftime('%Y-%m-%d'),
            "amount": row.amount
        }
        for row in employee_data
    ]

    return jsonify(result), 200

@app.route('/category-details', methods=['GET'])
def get_category_details():
    projectname = request.args.get('projectname')
    if not projectname:
        return jsonify({"error": "Missing projectname"}), 400

    category_data = (
        db.session.query(payments1.type, payments1.estamount, payments1.expense, options.material)
       # .join(options, payments1.type == options.labour)
        .filter(payments1.projectname == projectname)
        .all()
    )

    result = [
        {
            "type": row.type,
            "estamount": row.estamount,
            "expense": row.expense,
           # "category": row.material
        }
        for row in category_data
    ]

    return jsonify(result), 200

@app.route('/get_categories', methods=['GET'])
def get_categories():
    selected_type = request.args.get('type')

    if not selected_type:
        return jsonify({"error": "Missing type parameter"}), 400

    try:
        # Fetch all categories where type matches
        results = category.query.filter_by(type=selected_type).all()

        if not results:
            return jsonify({"error": "No matching categories found"}), 404

        # Convert results to JSON format
        categories = [{"category_id": row.id, "category_name": row.name} for row in results]

        return jsonify({"categories": categories})

    except Exception as e:
        print(f"Error: {str(e)}")
        return jsonify({"error": "Internal server error"}), 500
    
@app.route('/get_projects_by_id', methods=['GET'])
def get_projects_by_id():
    user_id = request.args.get('id')  # Get ID from the request
    if not user_id:
        return jsonify({"error": "ID parameter is required"}), 400

    projects = db.session.query(
        payments2.projectname, payments2.date, payments2.amount
    ).filter_by(id=user_id).all()
    
    project_details = [{
        "project_name": proj[0],
        "date": proj[1],
        "expense": proj[2]  # Assuming 'amount' represents the expense
    } for proj in projects]

    return jsonify({"projects": project_details})


@app.route('/check-overrun', methods=['POST'])
def check_overrun():
    data = request.get_json()
    projectname = data.get('projectname')

    if not projectname:
        return jsonify({"error": "Missing projectname"}), 400

    try:
        # Fetch project details
        project = projects.query.filter_by(projectname=projectname).first()
        if not project:
            return jsonify({"error": "Project not found"}), 404

        # Calculate duration and cost per day
        days_passed = (datetime.today().date() - project.startdate).days
        days_passed = max(days_passed, 1)  # Prevent division by zero
        
        cost_per_day = project.totexpense / days_passed

        project_progress = (days_passed / max(project.duration, 1)) * 100

        # ✅ Fetch dynamic types from options table
        labour_types = [opt.labour for opt in options.query.filter(options.labour != '').all()]
        material_types = [opt.material for opt in options.query.filter(options.material != '').all()]
        machinery_types = [opt.machinery for opt in options.query.filter(options.machinery != '').all()]

        # Fetch payments1 data for the project
        payments = payments1.query.filter_by(projectname=projectname).all()

        # ✅ Calculate costs dynamically
        labour_cost = sum(p.expense for p in payments if p.type in labour_types)
        material_cost = sum(p.expense for p in payments if p.type in material_types)
        machinery_cost = sum(p.expense for p in payments if p.type in machinery_types)

        # Prepare input for the model
        input_values = np.array([[
            project.quotedamount,  # Quoted Budget
            labour_cost,            # Labour Cost
            material_cost,          # Material Cost
            machinery_cost,         # Machinery Cost
            project_progress,       # Project Progress (%)
            cost_per_day            # Cost Per Day
        ]])

        # Convert to DataFrame
        columns = ['Quoted Budget', 'Labor Cost', 'Material Cost', 'Machinery Cost', 'Project Progress (%)', 'Cost Per Day']
        input_df = pd.DataFrame(input_values, columns=columns)

        # Make prediction
        predicted_cost = model.predict(input_df)[0]

      # Check for overrun (Convert to Python bool explicitly)
        overrun = bool(predicted_cost > project.quotedamount)

        return jsonify({
            "overrun": overrun,  # Now it's a Python bool, avoiding serialization issues
            "predicted_cost": float(predicted_cost)  # Also ensure predicted cost is float
        })


    except Exception as e:
        return jsonify({"error": str(e)}), 500
    
    # Update project details
@app.route('/update-project', methods=['POST'])
def update_project():
    data = request.get_json()
    projectname = data.get('projectname')
    quotedamount = data.get('quotedamount')
    totexpense = data.get('totexpense')

    if not projectname or quotedamount is None:
        return jsonify({"error": "Missing required fields"}), 400

    try:
        project = projects.query.filter_by(projectname=projectname).first()
        if not project:
            return jsonify({"error": "Project not found"}), 404

        project.quotedamount = quotedamount
        project.totexpense = totexpense
        db.session.commit()

        return jsonify({"message": "Project updated successfully"}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

@app.route('/delete-expense', methods=['POST'])
def delete_expense():
    data = request.get_json()
    
    # Validate required fields
    required_fields = ['uid']
    missing_fields = [field for field in required_fields if field not in data]
    
    if missing_fields:
        return jsonify({
            "error": "Missing required fields",
            "missing": missing_fields
        }), 400

    try:
        # Find the expense record in payments2 using uid
        payment = payments2.query.filter_by(uid=data['uid']).first()

        if not payment:
            return jsonify({
                "error": "Expense record not found",
                "details": f"No record found with uid={data['uid']}"
            }), 404

        amount = payment.amount
        projectname = payment.projectname
        category_id = payment.id  # id in payments2 table

        # Fetch project details and update total expense
        project = projects.query.filter_by(projectname=projectname).first()
        if project:
            project.totexpense -= amount  # Subtract from total expense

        # Fetch type from category table using category_id
        category_record = category.query.get(category_id)
        if not category_record:
            return jsonify({"error": "Invalid category ID"}), 400

        category_type = category_record.type  # Get type

        # Find the corresponding record in payments1 using projectname & type
        payment1 = payments1.query.filter_by(
            projectname=projectname,
            type=category_type
        ).first()

        if payment1:
            payment1.expense -= amount  # Subtract expense

        # Delete the expense from payments2
        db.session.delete(payment)

        db.session.commit()

        return jsonify({
            "message": "Expense deleted successfully",
            "deleted_amount": amount,
            "deleted_type": category_type
        }), 200

    except Exception as e:
        db.session.rollback()
        print("Delete Expense Error:", str(e))  # Debugging log
        return jsonify({
            "error": "Internal server error",
            "details": str(e)
        }), 500

# Add new payment type
@app.route('/add-payment-type', methods=['POST'])
def add_payment_type():
    data = request.get_json()
    projectname = data.get('projectname')
    type_ = data.get('type')
    estamount = data.get('estamount')
    expense = data.get('expense', 0)

    if not all([projectname, type_, estamount is not None]):
        return jsonify({"error": "Missing required fields"}), 400

    try:
        # Check if payment type already exists
        existing = payments1.query.filter_by(
            projectname=projectname,
            type=type_
        ).first()
        
        if existing:
            return jsonify({"error": "Payment type already exists"}), 400

        # Add new payment type
        new_payment = payments1(
            projectname=projectname,
            type=type_,
            estamount=estamount,
            expense=expense
        )
        db.session.add(new_payment)
        db.session.commit()
        
        return jsonify({"message": "Payment type added successfully"}), 201
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

# Update payment type
@app.route('/update-payment-type', methods=['POST'])
def update_payment_type():
    data = request.get_json()
    projectname = data.get('projectname')
    type_ = data.get('type')
    estamount = data.get('estamount')

    if not all([projectname, type_, estamount is not None]):
        return jsonify({"error": "Missing required fields"}), 400

    try:
        payment = payments1.query.filter_by(
            projectname=projectname,
            type=type_
        ).first()
        
        if not payment:
            return jsonify({"error": "Payment type not found"}), 404

        payment.estamount = estamount
        db.session.commit()
        
        return jsonify({"message": "Payment type updated successfully"}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500
    
@app.route('/get-expenses-for-deletion', methods=['GET'])
def get_expenses_for_deletion():
    projectname = request.args.get('projectname')
    if not projectname:
        return jsonify({"error": "Missing projectname"}), 400

    try:
        # Get all expenses with all fields needed for deletion
        expenses = db.session.query(
            payments2.uid,           # Primary key
            payments2.projectname,
            payments2.id,             # Category ID
            payments2.date,
            payments2.amount,
            category.name,           # Employee/category name
            category.type            # Expense type
        ).join(
            category, payments2.id == category.id
        ).filter(
            payments2.projectname == projectname
        ).all()

        result = [{
            "uid": exp.uid,
            "projectname": exp.projectname,
            "id": exp.id,
            "date": exp.date.strftime('%Y-%m-%d'),
            "amount": exp.amount,
            "name": exp.name,
            "type": exp.type
        } for exp in expenses]

        return jsonify(result), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500

# Report Generation Functions
def generate_pdf_report(data, title, filename):
    """Generate PDF report using ReportLab"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1  # Center alignment
    )
    
    # Add title
    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 20))
    
    # Add timestamp
    timestamp = Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal'])
    elements.append(timestamp)
    elements.append(Spacer(1, 20))
    
    if data:
        # Convert data to table format
        if isinstance(data, list) and len(data) > 0:
            # Get headers from first row
            headers = list(data[0].keys())
            table_data = [headers]  # Add headers as first row
            
            # Add data rows
            for row in data:
                table_data.append([str(row.get(header, '')) for header in headers])
            
            # Create table
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            elements.append(table)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    return buffer

def generate_excel_report(data, title, filename):
    """Generate Excel report using OpenPyXL"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    
    # Add title
    ws['A1'] = title
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:E1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Add timestamp
    ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A2'].font = Font(size=10, italic=True)
    ws.merge_cells('A2:E2')
    
    if data:
        # Add headers
        if isinstance(data, list) and len(data) > 0:
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            
            # Add data
            for row_idx, row_data in enumerate(data, 5):
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=str(row_data.get(header, '')))
                    cell.alignment = Alignment(horizontal='center')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

@app.route('/download-project-report-pdf', methods=['GET'])
def download_project_report_pdf():
    """Download project details report as PDF"""
    projectname = request.args.get('projectname')
    report_type = request.args.get('type', 'all')  # all, employee, category
    
    if not projectname:
        return jsonify({"error": "Missing projectname"}), 400
    
    try:
        # Get project details
        project = projects.query.filter_by(projectname=projectname).first()
        if not project:
            return jsonify({"error": "Project not found"}), 404
        
        data = []
        title = f"Project Report: {projectname}"
        
        if report_type == 'all' or report_type == 'employee':
            # Get employee details
            employee_data = (
                db.session.query(category.name, category.type, payments2.date, payments2.amount)
                .join(payments2, category.id == payments2.id)
                .filter(payments2.projectname == projectname)
                .all()
            )
            
            for row in employee_data:
                data.append({
                    "Name": row.name,
                    "Type": row.type,
                    "Date": row.date.strftime('%Y-%m-%d'),
                    "Amount": f"₹{row.amount:,}"
                })
        
        elif report_type == 'category':
            # Get category details
            category_data = (
                db.session.query(payments1.type, payments1.estamount, payments1.expense)
                .filter(payments1.projectname == projectname)
                .all()
            )
            
            for row in category_data:
                data.append({
                    "Type": row.type,
                    "Estimated Amount": f"₹{row.estamount:,}",
                    "Actual Expense": f"₹{row.expense:,}",
                    "Difference": f"₹{row.estamount - row.expense:,}"
                })
        
        # Generate PDF
        buffer = generate_pdf_report(data, title, f"{projectname}_report.pdf")
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f"{projectname}_report.pdf",
            mimetype='application/pdf'
        )
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/download-project-report-excel', methods=['GET'])
def download_project_report_excel():
    """Download project details report as Excel"""
    projectname = request.args.get('projectname')
    report_type = request.args.get('type', 'all')  # all, employee, category
    
    if not projectname:
        return jsonify({"error": "Missing projectname"}), 400
    
    try:
        # Get project details
        project = projects.query.filter_by(projectname=projectname).first()
        if not project:
            return jsonify({"error": "Project not found"}), 404
        
        data = []
        title = f"Project Report: {projectname}"
        
        if report_type == 'all' or report_type == 'employee':
            # Get employee details
            employee_data = (
                db.session.query(category.name, category.type, payments2.date, payments2.amount)
                .join(payments2, category.id == payments2.id)
                .filter(payments2.projectname == projectname)
                .all()
            )
            
            for row in employee_data:
                data.append({
                    "Name": row.name,
                    "Type": row.type,
                    "Date": row.date.strftime('%Y-%m-%d'),
                    "Amount": row.amount
                })
        
        elif report_type == 'category':
            # Get category details
            category_data = (
                db.session.query(payments1.type, payments1.estamount, payments1.expense)
                .filter(payments1.projectname == projectname)
                .all()
            )
            
            for row in category_data:
                data.append({
                    "Type": row.type,
                    "Estimated Amount": row.estamount,
                    "Actual Expense": row.expense,
                    "Difference": row.estamount - row.expense
                })
        
        # Generate Excel
        buffer = generate_excel_report(data, title, f"{projectname}_report.xlsx")
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f"{projectname}_report.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/download-user-financial-report-pdf', methods=['GET'])
def download_user_financial_report_pdf():
    """Download user financial details report as PDF"""
    user_id = request.args.get('id')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    if not user_id:
        return jsonify({"error": "Missing user ID"}), 400
    
    try:
        # Build query
        query = db.session.query(
            payments2.projectname, payments2.date, payments2.amount,
            category.name, category.type
        ).join(category, payments2.id == category.id).filter(payments2.id == user_id)
        
        # Apply date filters if provided
        if start_date:
            query = query.filter(payments2.date >= start_date)
        if end_date:
            query = query.filter(payments2.date <= end_date)
        
        financial_data = query.all()
        
        data = []
        for row in financial_data:
            data.append({
                "Project Name": row.projectname,
                "Date": row.date.strftime('%Y-%m-%d'),
                "Amount": f"₹{row.amount:,}",
                "Category": row.name,
                "Type": row.type
            })
        
        title = f"User Financial Report (ID: {user_id})"
        if start_date and end_date:
            title += f" - {start_date} to {end_date}"
        
        # Generate PDF
        buffer = generate_pdf_report(data, title, f"user_{user_id}_financial_report.pdf")
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f"user_{user_id}_financial_report.pdf",
            mimetype='application/pdf'
        )
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/download-user-financial-report-excel', methods=['GET'])
def download_user_financial_report_excel():
    """Download user financial details report as Excel"""
    user_id = request.args.get('id')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    if not user_id:
        return jsonify({"error": "Missing user ID"}), 400
    
    try:
        # Build query
        query = db.session.query(
            payments2.projectname, payments2.date, payments2.amount,
            category.name, category.type
        ).join(category, payments2.id == category.id).filter(payments2.id == user_id)
        
        # Apply date filters if provided
        if start_date:
            query = query.filter(payments2.date >= start_date)
        if end_date:
            query = query.filter(payments2.date <= end_date)
        
        financial_data = query.all()
        
        data = []
        for row in financial_data:
            data.append({
                "Project Name": row.projectname,
                "Date": row.date.strftime('%Y-%m-%d'),
                "Amount": row.amount,
                "Category": row.name,
                "Type": row.type
            })
        
        title = f"User Financial Report (ID: {user_id})"
        if start_date and end_date:
            title += f" - {start_date} to {end_date}"
        
        # Generate Excel
        buffer = generate_excel_report(data, title, f"user_{user_id}_financial_report.xlsx")
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f"user_{user_id}_financial_report.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True,host="0.0.0.0",port=5000)
